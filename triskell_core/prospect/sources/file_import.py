"""
Import de prospects depuis un fichier utilisateur (CSV, XLSX).

Workflow :
  1. `preview(path)` → renvoie (headers, rows_sample, suggested_mapping)
  2. l'UI laisse l'utilisateur ajuster `mapping`
  3. `import_with_mapping(path, mapping)` → yield des Prospect

Le mapping est `{champ_prospect: nom_colonne_fichier}`, ex :
  {"name": "Nom", "emails": "Email", "city": "Ville"}

Aucune dépendance externe pour CSV. Pour XLSX, dépend de openpyxl.
"""

from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path
from typing import Iterator

from ..core.prospect import Prospect, Source


# Champs prospect qu'on sait importer depuis un fichier
IMPORTABLE_FIELDS = [
    ("name",        "Nom"),
    ("legal_name",  "Raison sociale"),
    ("emails",      "Email"),
    ("phones",      "Téléphone"),
    ("website",     "Site web"),
    ("address",     "Adresse"),
    ("city",        "Ville"),
    ("postal_code", "Code postal"),
    ("country",     "Pays"),
    ("industry",    "Secteur"),
    ("siren",       "SIREN"),
    ("notes",       "Notes"),
]

# Synonymes courants pour l'auto-détection (header normalisé → champ prospect)
_HEADER_HINTS: dict[str, list[str]] = {
    "name": [
        "nom", "name", "nomcomplet", "fullname", "contact", "prenomnom",
        "prenom", "firstname", "nomprenom", "intitule",
    ],
    "legal_name": [
        "raisonsociale", "societe", "entreprise", "company", "companyname",
        "denomination", "etablissement",
    ],
    "emails": [
        "email", "mail", "emails", "mails", "courriel", "adressemail",
        "adresseemail", "emailaddress", "emailpro",
    ],
    "phones": [
        "tel", "telephone", "phone", "phonenumber", "mobile", "portable",
        "telmobile", "telpro", "numero", "numerotelephone",
    ],
    "website": [
        "site", "siteweb", "website", "url", "web", "lien", "internet",
        "siteinternet",
    ],
    "address": [
        "adresse", "address", "rue", "voie", "addressline",
    ],
    "city": [
        "ville", "city", "town", "commune", "localite",
    ],
    "postal_code": [
        "cp", "codepostal", "postal", "zip", "zipcode", "postalcode",
    ],
    "country": [
        "pays", "country", "nation",
    ],
    "industry": [
        "secteur", "industry", "industrie", "metier", "domaine", "activite",
        "categorie", "type", "naf",
    ],
    "siren": [
        "siren", "siret", "rcs", "numerosiren", "numerosiret",
    ],
    "notes": [
        "notes", "note", "commentaire", "commentaires", "remarque",
        "remarques", "observation",
    ],
}


# ---------------------------------------------------------------------------
# Normalisation des en-têtes pour matching
# ---------------------------------------------------------------------------
def _norm_header(h: str) -> str:
    """Minuscule, sans accents, sans ponctuation/espaces."""
    if not h:
        return ""
    nfkd = unicodedata.normalize("NFKD", h)
    no_acc = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]", "", no_acc.lower())


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    """Devine quelle colonne correspond à quel champ prospect."""
    norm_to_orig = {_norm_header(h): h for h in headers if h}
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for field, _label in IMPORTABLE_FIELDS:
        hints = _HEADER_HINTS.get(field, [])
        for hint in hints:
            if hint in norm_to_orig and norm_to_orig[hint] not in used:
                mapping[field] = norm_to_orig[hint]
                used.add(norm_to_orig[hint])
                break
        if field not in mapping:
            # fallback : un header qui contient le hint
            for norm, orig in norm_to_orig.items():
                if orig in used:
                    continue
                if any(hint in norm for hint in hints):
                    mapping[field] = orig
                    used.add(orig)
                    break
    return mapping


# ---------------------------------------------------------------------------
# Lecture brute
# ---------------------------------------------------------------------------
def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    # Auto-détection du séparateur (virgule, ;, tab)
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    sniffer = csv.Sniffer()
    sample = raw[:4096]
    try:
        dialect = sniffer.sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class _D:
            delimiter = ","
            quotechar = '"'
        dialect = _D()  # type: ignore
    reader = csv.DictReader(raw.splitlines(), dialect=dialect)
    rows = [dict(r) for r in reader]
    headers = list(reader.fieldnames or [])
    return headers, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "Lecture Excel impossible : module openpyxl manquant. "
            "Installer avec : pip install openpyxl"
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows: list[dict[str, str]] = []
    for row in rows_iter:
        if row is None:
            continue
        d: dict[str, str] = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            key = headers[i]
            if not key:
                continue
            if val is None:
                d[key] = ""
            elif isinstance(val, float) and val.is_integer():
                d[key] = str(int(val))
            else:
                d[key] = str(val).strip()
        if any(v for v in d.values()):
            rows.append(d)
    return headers, rows


def read_file(path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    """Lit un CSV/XLSX → (headers, rows). Rows = liste de dict {colonne: valeur}."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    suffix = p.suffix.lower()
    if suffix in (".csv", ".tsv", ".txt"):
        return _read_csv(p)
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return _read_xlsx(p)
    raise ValueError(f"Format de fichier non supporté : {suffix}")


def preview(path: str | Path, *, sample_size: int = 5) -> dict:
    """Renvoie un aperçu du fichier pour l'UI de mapping."""
    headers, rows = read_file(path)
    return {
        "headers": headers,
        "rows_total": len(rows),
        "sample": rows[:sample_size],
        "suggested_mapping": suggest_mapping(headers),
    }


# ---------------------------------------------------------------------------
# Conversion en Prospect
# ---------------------------------------------------------------------------
_LIST_SPLIT_RE = re.compile(r"[,;|/]")


def _split_list(value: str) -> list[str]:
    """Découpe une cellule contenant plusieurs valeurs (emails, tels)."""
    if not value:
        return []
    parts = [p.strip() for p in _LIST_SPLIT_RE.split(value) if p.strip()]
    return parts


def _row_to_prospect(row: dict[str, str], mapping: dict[str, str],
                     source_label: str) -> Prospect | None:
    """Convertit une ligne fichier → Prospect via le mapping."""
    def get(field: str) -> str:
        col = mapping.get(field)
        if not col:
            return ""
        return (row.get(col) or "").strip()

    name = get("name")
    legal_name = get("legal_name")
    emails = _split_list(get("emails"))
    phones = _split_list(get("phones"))
    website = get("website")
    siren = get("siren")

    # Au moins UNE info utile sinon on saute
    if not (name or legal_name or emails or phones or website or siren):
        return None

    p = Prospect(
        name=name,
        legal_name=legal_name,
        emails=emails,
        phones=phones,
        website=website,
        address=get("address"),
        city=get("city"),
        postal_code=get("postal_code"),
        country=get("country"),
        industry=get("industry"),
        siren=siren,
        notes=get("notes"),
        sources=[Source(name="file", source_id=source_label)],
    )
    return p


def import_with_mapping(path: str | Path, mapping: dict[str, str]) -> Iterator[Prospect]:
    """Itère les prospects du fichier convertis via le mapping fourni."""
    p = Path(path)
    _headers, rows = read_file(p)
    label = p.name
    for row in rows:
        prospect = _row_to_prospect(row, mapping, source_label=label)
        if prospect is not None:
            yield prospect
